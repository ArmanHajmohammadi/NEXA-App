from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication, QFileDialog, QGroupBox, QVBoxLayout, QHBoxLayout, QLineEdit, QLabel, QPushButton, QMainWindow, QSpinBox, QProgressBar
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from pathlib import Path
import os
import re
from PyQt5.QtGui import QIcon
import sys
from PyQt5.QtCore import Qt, QTimer
import sys

workers = []

path = Path('./report.xlsx')
if (path.is_file()):
    os.remove('./report.xlsx')

# Create a new workbook
workbook = openpyxl.Workbook()

# Save the workbook to the specified file path
workbook.save('./report.xlsx')

report = openpyxl.load_workbook('./report.xlsx')
reportSheet = report.active


class ReporterApp(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Reporter")
        self.setWindowIcon(QIcon("./Icon.ico"))
        self.init_ui()

    def init_ui(self):
        self.layout = QtWidgets.QVBoxLayout()

        # Create GroupBox 1 - Choose Excel File
        groupbox1 = QGroupBox("Choose Excel File")
        hbox1 = QHBoxLayout()
        self.file_lineedit = QLineEdit()
        self.choose_file_button = QPushButton("Browse")
        self.choose_file_button.clicked.connect(self.open_file_dialog)
        hbox1.addWidget(self.file_lineedit)
        hbox1.addWidget(self.choose_file_button)
        groupbox1.setLayout(hbox1)

        # Create GroupBox 2 - Spin Field, Label, and Text Fields
        groupbox2 = QGroupBox("Spin Field, Label, and Text Fields")
        vbox2 = QVBoxLayout()
        self.spin_field = QSpinBox()
        self.spin_field.setMinimum(0)
        self.spin_field.valueChanged.connect(self.generate_text_fields)
        vbox2.addWidget(self.spin_field)

        self.label = QLabel("Enter the title cells you wanna :")
        vbox2.addWidget(self.label)

        self.text_fields = []
        groupbox2.setLayout(vbox2)

        # Create Run Button
        self.run_button = QPushButton("Run")
        self.run_button.clicked.connect(self.run_report)

        # Create a progress bar
        self.progress_bar = QProgressBar()

        # Set the minimum and maximum values for the progress bar
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)

        # Add GroupBoxes and Run Button to the main layout
        self.layout.addWidget(groupbox1)
        self.layout.addWidget(groupbox2)
        self.layout.addWidget(self.run_button)
        self.layout.addWidget(self.progress_bar, stretch=1)

        # Set main layout for the window
        self.setLayout(self.layout)

    def open_file_dialog(self):
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getOpenFileName(
            self, "Select Excel File", filter="Excel Files (*.xlsx *.xls)")
        self.file_lineedit.setText(file_path)

    def generate_text_fields(self):
        num_fields = self.spin_field.value()
        vbox2 = self.layout.itemAt(1).widget().layout()

        while vbox2.count() > 2:
            item = vbox2.takeAt(2)
            if item.widget():
                item.widget().deleteLater()

        self.text_fields = []
        for i in range(num_fields):
            text_field = QLineEdit()
            text_field.setPlaceholderText(
                "The first cell of the column (E.g. 'A1' )")
            vbox2.addWidget(text_field)
            self.text_fields.append(text_field)

    def update_progress(self, value):
        self.progress_bar.setValue(value)

    def run_report(self):
        self.run_button.setText("Processing...")
        self.run_button.setEnabled(False)
        file_path = self.file_lineedit.text()
        num_fields = self.spin_field.value()
        text_values = [field.text() for field in self.text_fields]
        print("Excel File Path:", file_path)
        print("Number of Fields:", num_fields)
        print("Text Values:", text_values)
        for i in range(1, 32):

            reportSheet.cell(
                row=1, column=i*(len(text_values)-1)).value = str(i)
            temple_path = ""
            if (i < 10):
                temple_path = file_path
                temple_path = temple_path.replace("01", "0" + str(i))
                path = Path(temple_path)
                address = temple_path
            else:
                temple_path = file_path
                temple_path = temple_path.replace("01", str(i))
                path = Path(temple_path)
                address = temple_path

            if (path.is_file()):
                wb = openpyxl.load_workbook(address)
                sh = wb.active

                # print(intendedCell.value)
                row = 1
                intendedCell = sh[text_values[0]]
                while (intendedCell.value != None):
                    if (not (intendedCell.value in workers)):
                        workers.append(intendedCell.value)

                    for j in range(1, len(text_values)):
                        match = re.match(
                            r"([a-zA-Z]+)([0-9]+)", text_values[j])

                        reportSheet.cell(row=workers.index(intendedCell.value)+2, column=j+i*(len(text_values)-1)-1).value = sh[
                            match.group(1) + str(row + int(match.group(2))-1)].value

                    row += 1
                    match = re.match(
                        r"([a-zA-Z]+)([0-9]+)", text_values[0])
                    intendedCell = sh[match.group(
                        1) + str(row + int(match.group(2)))]
            self.update_progress(int(i/31 * 100))

        print("******* Completed *******")

        for i in range(0, len(workers)):
            reportSheet.cell(row=i+2, column=1).value = workers[i]
        reportSheet.cell(row=1, column=1).value = "روز"
        reportSheet.sheet_view.rightToLeft = True
        report.save('./report.xlsx')
        self.run_button.setText("Completed!")
        print("Saved")


if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    window = ReporterApp()
    window.show()
    sys.exit(app.exec_())
